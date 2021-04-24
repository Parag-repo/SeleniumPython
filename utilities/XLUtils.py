import openpyxl
from openpyxl import workbook


def getRowCount(file, sheetName):
    workbook1 = openpyxl.load_workbook(file)
    sheet = workbook1[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook1 = openpyxl.load_workbook(file)
    sheet = workbook1[sheetName]
    return sheet.max_column


def readData(file, sheetName, rownum, columnno):
    workbook1 = openpyxl.load_workbook(file)
    sheet = workbook1[sheetName]
    return sheet.cell(row=rownum, column=columnno).value


def writeData(file, sheetName, rownum, columnno, data):
    workbook1 = openpyxl.load_workbook(file)
    sheet = workbook1[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook1.save(file)
